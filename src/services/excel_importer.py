import csv
import io
from openpyxl import load_workbook
from storage.database import Driver, db


def _load_file(file_obj):
    """Load Excel or CSV file and return list of dictionaries"""
    filename = getattr(file_obj, 'filename', '') or ''
    lower_name = filename.lower()

    if lower_name.endswith('.xls'):
        raise ValueError("Legacy .xls files are not supported. Please upload .xlsx or .csv.")
    
    if lower_name.endswith('.csv'):
        # Handle CSV
        content = file_obj.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        return list(reader)
    else:
        # Handle Excel
        wb = load_workbook(file_obj)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Convert to list of dictionaries
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return data


def _resolve_column(data, accepted_names):
    """Find column name that matches accepted names (case-insensitive)"""
    if not data:
        return None
    
    first_row = data[0]
    for column_name in first_row.keys():
        if column_name:
            normalized = str(column_name).strip().lower()
            if normalized in accepted_names:
                return column_name
    return None


def import_excel(file_obj):
    """Import DA list (Excel or CSV) and merge by Transporter ID, preserving existing non-compliance counts."""
    data = _load_file(file_obj)

    transport_column = _resolve_column(data, {"transporter id"})
    name_column = _resolve_column(data, {"driver name", "name"})

    if not transport_column or not name_column:
        raise ValueError("File must include columns: 'Transporter ID' and 'Driver Name'.")

    added = 0
    updated = 0
    skipped = 0

    for row in data:
        transport_id = str(row.get(transport_column, "")).strip()
        driver_name = str(row.get(name_column, "")).strip()

        if not transport_id or transport_id.lower() == "nan" or not driver_name or driver_name.lower() == "nan":
            skipped += 1
            continue

        driver = Driver.query.filter_by(transport_id=transport_id).first()
        if driver is None:
            db.session.add(Driver(transport_id=transport_id, name=driver_name))
            added += 1
        else:
            if driver.name != driver_name:
                driver.name = driver_name
                updated += 1

    db.session.commit()

    return {
        "total_rows": len(data),
        "added": added,
        "updated": updated,
        "skipped": skipped,
    }