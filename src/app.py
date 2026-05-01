from datetime import datetime
from io import BytesIO
import os

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, url_for

from services.excel_importer import import_excel
from services.non_compliance_manager import NonComplianceManager
from services.week_utils import get_operational_week
from storage.database import Driver, Incident, BikeIncident, init_db, db

app = Flask(__name__)
app.secret_key = "local-incident-tracker-secret"

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "incident_tracker.db")
if not os.path.exists(DB_PATH) and os.path.isdir(DATA_DIR):
    legacy_candidates = [name for name in os.listdir(DATA_DIR) if name.endswith(".db")]
    if legacy_candidates:
        DB_PATH = os.path.join(DATA_DIR, legacy_candidates[0])
db_url = os.environ.get('DATABASE_URL')
if db_url:
    # Fix postgres:// → postgresql:// for SQLAlchemy
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = db_url
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

non_compliance_manager = NonComplianceManager()

@app.route('/', methods=['GET', 'POST'])
def index():
    today = datetime.today().date()
    current_week = get_operational_week(today)
    drivers = Driver.query.order_by(Driver.name.asc()).all()
    escalations = non_compliance_manager.get_escalation_drivers()
    return render_template(
        'index.html',
        drivers=drivers,
        escalations=escalations,
        today=today.isoformat(),
        current_week=current_week,
    )


@app.route('/drivers.json')
def drivers_json():
    try:
        drivers = Driver.query.order_by(Driver.name.asc()).all()
        return jsonify([{"transport_id": d.transport_id, "name": d.name} for d in drivers])
    except Exception:
        app.logger.exception('Failed to load drivers for drivers.json')
        return jsonify([])


@app.route('/upload-da-list', methods=['POST'])
def upload_da_list():
    file = request.files.get('da_file')
    if not file or file.filename == '':
        flash('Please select an Excel file to upload.', 'error')
        return redirect(url_for('index'))

    try:
        summary = import_excel(file)
        flash(
            f"DA list imported. Rows: {summary['total_rows']}, added: {summary['added']}, updated: {summary['updated']}, skipped: {summary['skipped']}.",
            'success',
        )
    except Exception as exc:
        flash(f'Import failed: {exc}', 'error')

    return redirect(url_for('index'))


@app.route('/add-da', methods=['POST'])
def add_da():
    transport_id = request.form.get('new_transport_id', '').strip()
    driver_name = request.form.get('new_driver_name', '').strip()

    if not transport_id or not driver_name:
        flash('Transporter ID and Driver Name are required.', 'error')
        return redirect(url_for('index'))

    existing = Driver.query.filter_by(transport_id=transport_id).first()
    if existing:
        flash(f'Driver with Transporter ID {transport_id} already exists.', 'error')
        return redirect(url_for('index'))

    try:
        db.session.add(Driver(transport_id=transport_id, name=driver_name))
        db.session.commit()
        flash(f'Driver {driver_name} ({transport_id}) added.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f'Failed to add driver: {exc}', 'error')

    return redirect(url_for('index'))


@app.route('/clear-week-incidents', methods=['POST'])
def clear_week_incidents():
    from datetime import timedelta
    
    current_week = get_operational_week(datetime.today().date())
    week_start = current_week["week_start"]
    week_end = current_week["week_end"]
    
    try:
        incidents_to_delete = Incident.query.filter(
            Incident.incident_date >= week_start,
            Incident.incident_date <= week_end,
        ).all()
        
        count = len(incidents_to_delete)
        for incident in incidents_to_delete:
            db.session.delete(incident)
        
        # Reset non_compliance_count for drivers affected this week
        for incident in incidents_to_delete:
            driver = Driver.query.get(incident.driver_id)
            if driver and driver.non_compliance_count > 0:
                driver.non_compliance_count -= 1
        
        db.session.commit()
        flash(f'Cleared {count} incident(s) for the current week.', 'success')
    except Exception as exc:
        db.session.rollback()
        flash(f'Failed to clear incidents: {exc}', 'error')
    
    return redirect(url_for('index'))


@app.route('/record-incident', methods=['POST'])
def record_incident():
    transport_id = request.form.get('transport_id', '').strip()
    missing_vest = request.form.get('missing_vest') == 'on'
    missing_shoes = request.form.get('missing_shoes') == 'on'
    no_badge = request.form.get('no_badge') == 'on'
    not_following_yard_marshall_instructions = request.form.get('not_following_yard_marshall_instructions') == 'on'
    exceeding_speed_loading_bay = request.form.get('exceeding_speed_loading_bay') == 'on'
    unnecessary_dwell_time = request.form.get('unnecessary_dwell_time') == 'on'
    incident_date_raw = request.form.get('incident_date', '').strip()

    if not transport_id:
        flash('Please select a driver from the dropdown.', 'error')
        return redirect(url_for('index'))

    try:
        incident_date = datetime.strptime(incident_date_raw, '%Y-%m-%d').date()
    except ValueError:
        flash('Please provide a valid incident date.', 'error')
        return redirect(url_for('index'))

    result = non_compliance_manager.record_incident(
        transporter_id=transport_id or None,
        missing_vest=missing_vest,
        missing_shoes=missing_shoes,
        no_badge=no_badge,
        not_following_yard_marshall_instructions=not_following_yard_marshall_instructions,
        exceeding_speed_loading_bay=exceeding_speed_loading_bay,
        unnecessary_dwell_time=unnecessary_dwell_time,
        incident_date=incident_date,
    )

    flash(result['message'], 'success' if result['ok'] else 'error')
    return redirect(url_for('index'))


@app.route('/escalations.xlsx', methods=['GET'])
def export_escalations_csv():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import timedelta

    # Get week offset from query parameter (default 0 for current week, -1 for past week, etc.)
    week_offset = request.args.get('week_offset', 0, type=int)
    target_date = datetime.today().date()
    if week_offset != 0:
        # Move target_date by the offset (7 days per week)
        target_date = target_date + timedelta(weeks=week_offset)
    
    current_week = get_operational_week(target_date)
    week_start = current_week["week_start"]
    week_end = current_week["week_end"]

    weekly_incidents = Incident.query.filter(
        Incident.incident_date >= week_start,
        Incident.incident_date <= week_end,
    ).all()
    driver_ids = sorted({incident.driver_id for incident in weekly_incidents})
    drivers = Driver.query.filter(Driver.id.in_(driver_ids)).order_by(Driver.name.asc()).all() if driver_ids else []

    def incident_count(driver_id, column):
        return Incident.query.filter(
            Incident.driver_id == driver_id,
            Incident.incident_date >= week_start,
            Incident.incident_date <= week_end,
            column == True
        ).count()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {current_week['week_number']}"

    headers = [
        'Transporter ID',
        'Driver Name',
        'Non-Compliance Count (Weekly)',
        'Missing High Visibility Vest (Weekly)',
        'Missing Safety Shoes (Weekly)',
        'Not Displaying Badge (Weekly)',
        'Not Following Yard Marshall Instructions (Weekly)',
        'Exceeding 5mph in Loading Bay (Weekly)',
        'Unnecessary Dwell Time (Weekly)',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 36

    for driver in drivers:
        vest_count  = incident_count(driver.id, Incident.high_visibility_vest_missing)
        shoes_count = incident_count(driver.id, Incident.safety_shoes_missing)
        badge_count = incident_count(driver.id, Incident.no_badge)
        yard_count  = incident_count(driver.id, Incident.not_following_yard_marshall_instructions)
        speed_count = incident_count(driver.id, Incident.exceeding_speed_loading_bay)
        dwell_count = incident_count(driver.id, Incident.unnecessary_dwell_time)

        weekly_total = (
            vest_count + shoes_count + badge_count + yard_count + speed_count + dwell_count
        )

        # Non-Compliance Count cell: append escalation text in the same box when threshold reached
        non_compliance_cell_value = weekly_total
        if weekly_total >= 3:
            non_compliance_cell_value = f"{weekly_total} — escalation required"

        ws.append([
            driver.transport_id,
            driver.name,
            non_compliance_cell_value,
            vest_count,
            shoes_count,
            badge_count,
            yard_count,
            speed_count,
            dwell_count,
        ])

    # Auto-fit column widths based on the widest cell value
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    # Apply red fill to Non-Compliance Count cells that include escalation text
    red_fill = PatternFill(fill_type='solid', fgColor='FF0000')
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value and isinstance(cell.value, str) and 'escalation required' in cell.value.lower():
            cell.fill = red_fill
            cell.font = Font(bold=True, color='FFFFFF')

    bytes_buffer = BytesIO()
    wb.save(bytes_buffer)
    bytes_buffer.seek(0)
    return send_file(
        bytes_buffer,
        as_attachment=True,
        download_name=(
            f"incident_tracker_week_{current_week['week_number']}_"
            f"{week_start.isoformat()}_to_{week_end.isoformat()}.xlsx"
        ),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/bikes', methods=['GET', 'POST'])
def bikes():
    """Display bike incident recording form or handle clear action"""
    # Handle clearing bike incidents for the week
    if request.method == 'POST' and request.form.get('action') == 'clear':
        try:
            today = datetime.today().date()
            current_week = get_operational_week(today)
            week_start = current_week['week_start']
            week_end = current_week['week_end']

            # Delete all bike incidents for this week
            deleted_count = BikeIncident.query.filter(
                BikeIncident.incident_date >= week_start,
                BikeIncident.incident_date <= week_end
            ).delete()
            db.session.commit()

            if deleted_count > 0:
                flash(f'Cleared {deleted_count} bike incident(s) for this week', 'success')
            else:
                flash('No bike incidents to clear for this week', 'success')
            
        except Exception as e:
            app.logger.exception('Error clearing week bike incidents')
            flash(f'Error clearing bike incidents: {str(e)}', 'error')
    
    today = datetime.today().date()
    current_week = get_operational_week(today)
    return render_template(
        'bikes.html',
        today=today.isoformat(),
        current_week=current_week,
    )


@app.route('/record-bike-incident', methods=['POST'])
def record_bike_incident():
    """Record a bike incident"""
    try:
        bike_type = request.form.get('bike_type', '').strip()
        bike_serial = request.form.get('bike_serial', '').strip()
        incident_date_str = request.form.get('incident_date', '')

        # Validation
        if not bike_type or bike_type not in ['mubea', 'citkar']:
            flash('Please select a valid bike type', 'error')
            return redirect(url_for('bikes'))

        if not bike_serial:
            flash('Bike serial number is required', 'error')
            return redirect(url_for('bikes'))

        if not incident_date_str:
            flash('Incident date is required', 'error')
            return redirect(url_for('bikes'))

        incident_date = datetime.fromisoformat(incident_date_str).date()

        # Check at least one issue is selected
        issues = {
            'roll_back_on_ramp': request.form.get('roll_back_on_ramp') == 'on',
            'missing_windscreen': request.form.get('missing_windscreen') == 'on',
            'faulty_handbrake': request.form.get('faulty_handbrake') == 'on',
            'faulty_brake': request.form.get('faulty_brake') == 'on',
            'shutting_off': request.form.get('shutting_off') == 'on',
            'worn_out_tyres': request.form.get('worn_out_tyres') == 'on',
        }

        if not any(issues.values()):
            flash('Please select at least one bike issue', 'error')
            return redirect(url_for('bikes'))

        # Create and save bike incident
        bike_incident = BikeIncident(
            bike_type=bike_type,
            bike_serial=bike_serial,
            incident_date=incident_date,
            roll_back_on_ramp=issues['roll_back_on_ramp'],
            missing_windscreen=issues['missing_windscreen'],
            faulty_handbrake=issues['faulty_handbrake'],
            faulty_brake=issues['faulty_brake'],
            shutting_off=issues['shutting_off'],
            worn_out_tyres=issues['worn_out_tyres'],
        )
        db.session.add(bike_incident)
        db.session.commit()

        flash('Bike incident recorded successfully', 'success')
        return redirect(url_for('bikes'))

    except Exception as e:
        app.logger.exception('Error recording bike incident')
        flash(f'Error recording bike incident: {str(e)}', 'error')
        return redirect(url_for('bikes'))


@app.route('/bike-incidents.xlsx')
def export_bike_incidents():
    """Export bike incidents to Excel for a specific week"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        # Get week offset from query parameter (default 0 = current week, -1 = last week)
        week_offset = int(request.args.get('week_offset', 0))
        target_date = datetime.today().date()
        
        # Calculate target date for the requested week
        if week_offset != 0:
            target_date = target_date - __import__('datetime').timedelta(weeks=-week_offset)
        
        current_week = get_operational_week(target_date)
        week_start = current_week['week_start']
        week_end = current_week['week_end']

        # Get bike incidents for the specified week
        bike_incidents = BikeIncident.query.filter(
            BikeIncident.incident_date >= week_start,
            BikeIncident.incident_date <= week_end
        ).order_by(BikeIncident.incident_date.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"Week {current_week['week_number']}"

        headers = [
            'Bike Type',
            'Bike Serial',
            'Incident Date',
            'Roll Back on Ramp',
            'Missing Windscreen',
            'Faulty Handbrake',
            'Faulty Brake',
            'Shutting Off',
            'Worn Out Tyres',
            'Recorded At',
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1D4ED8')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 36

        # Add data rows
        for incident in bike_incidents:
            ws.append([
                incident.bike_type,
                incident.bike_serial,
                incident.incident_date.isoformat(),
                'Yes' if incident.roll_back_on_ramp else 'No',
                'Yes' if incident.missing_windscreen else 'No',
                'Yes' if incident.faulty_handbrake else 'No',
                'Yes' if incident.faulty_brake else 'No',
                'Yes' if incident.shutting_off else 'No',
                'Yes' if incident.worn_out_tyres else 'No',
                incident.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])

        # Auto-fit column widths
        for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
            max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

        bytes_buffer = BytesIO()
        wb.save(bytes_buffer)
        bytes_buffer.seek(0)
        return send_file(
            bytes_buffer,
            as_attachment=True,
            download_name=(
                f"bike_incidents_week_{current_week['week_number']}_"
                f"{week_start.isoformat()}_to_{week_end.isoformat()}.xlsx"
            ),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    except Exception as e:
        app.logger.exception('Error exporting bike incidents')
        flash(f'Error exporting bike incidents: {str(e)}', 'error')
        return redirect(url_for('bikes'))


@app.route('/clear-week-bike-incidents', methods=['POST'])
def clear_week_bike_incidents():
    """Delete all bike incidents for the current week"""
    try:
        today = datetime.today().date()
        current_week = get_operational_week(today)
        week_start = current_week['week_start']
        week_end = current_week['week_end']

        # Delete all bike incidents for this week
        deleted_count = BikeIncident.query.filter(
            BikeIncident.incident_date >= week_start,
            BikeIncident.incident_date <= week_end
        ).delete()
        db.session.commit()

        if deleted_count > 0:
            flash(f'Cleared {deleted_count} bike incident(s) for this week', 'success')
        else:
            flash('No bike incidents to clear for this week', 'success')
        
        return redirect(url_for('bikes'))

    except Exception as e:
        app.logger.exception('Error clearing week bike incidents')
        flash(f'Error clearing bike incidents: {str(e)}', 'error')
        return redirect(url_for('bikes'))


if __name__ == '__main__':
    os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)
    init_db(app)
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)